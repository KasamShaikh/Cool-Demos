"""Generate an Excel workbook from the latest FD-rate JSON in Blob Storage.

Each bank's data goes into a separate worksheet/tab.  The finished .xlsx
is uploaded back to the same Blob container and the blob URL is returned.
"""

import io
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.storage.blob_client import (
    get_blob_service_client,
    get_latest_from_blob,
    ensure_container_exists,
)

from azure.storage.blob import ContentSettings

# ── Excel styling constants ──────────────────────────────────────
_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="97144D", end_color="97144D", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_COLUMNS = [
    ("Category", 18),
    ("Tenor (Description)", 28),
    ("Min Days", 12),
    ("Max Days", 12),
    ("Rate %", 12),
    ("Amount Slab", 22),
    ("Scheme", 22),
    ("Effective Date", 16),
    ("Additional Info", 30),
]

_FIELD_MAP = [
    "category",
    "tenor_description",
    "tenor_min_days",
    "tenor_max_days",
    "rate_percent",
    "amount_slab",
    "scheme_name",
    "effective_date",
    "additional_info",
]


def _safe_sheet_name(name: str) -> str:
    """Sanitise a string so it can be used as an Excel sheet name."""
    for ch in r"[]:*?/\\":
        name = name.replace(ch, "")
    return name[:31]  # Excel limit


def generate_excel_bytes(data: dict) -> bytes:
    """Create an in-memory .xlsx workbook from consolidated FD data.

    Returns the raw bytes of the workbook.
    """
    wb = Workbook()
    # Remove the default sheet that openpyxl creates
    wb.remove(wb.active)

    banks = data.get("banks", [])
    if not banks:
        ws = wb.create_sheet("No Data")
        ws["A1"] = "No bank data found in the latest scrape results."
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    for bank in banks:
        bank_name = bank.get("bank_name", "Unknown")
        ws = wb.create_sheet(_safe_sheet_name(bank_name))

        # ── Title row ────────────────────────────────────────────
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_COLUMNS))
        title_cell = ws.cell(
            row=1, column=1, value=f"{bank_name} — Fixed Deposit Rates"
        )
        title_cell.font = Font(name="Calibri", bold=True, size=14, color="97144D")
        title_cell.alignment = Alignment(horizontal="center")

        # Meta row
        scraped_at = bank.get("scraped_at", "")
        source_url = bank.get("source_url", "")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(_COLUMNS))
        meta_cell = ws.cell(
            row=2, column=1, value=f"Source: {source_url}  |  Scraped: {scraped_at}"
        )
        meta_cell.font = Font(name="Calibri", size=9, color="666666")
        meta_cell.alignment = Alignment(horizontal="center")

        # ── Header row (row 4) ───────────────────────────────────
        header_row = 4
        for col_idx, (col_name, col_width) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _HEADER_ALIGN
            cell.border = _THIN_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        # ── Data rows ────────────────────────────────────────────
        rates = bank.get("rates", [])
        for row_offset, rate in enumerate(rates):
            row_num = header_row + 1 + row_offset
            for col_idx, field in enumerate(_FIELD_MAP, start=1):
                value = rate.get(field, "")
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.border = _THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Freeze panes below header
        ws.freeze_panes = f"A{header_row + 1}"
        # Auto-filter
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(len(_COLUMNS))}{header_row + len(rates)}"
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_excel_to_blob(container_name: str | None = None) -> dict:
    """Read latest.json, build Excel, upload to blob.

    Returns {"excel_url": "<blob url>", "file_name": "<name>"}.
    """
    container_name = container_name or os.getenv(
        "AZURE_STORAGE_CONTAINER_NAME", "fd-rates"
    )

    data = get_latest_from_blob(container_name)
    if data is None:
        raise ValueError(
            "No latest FD rate data found in Blob Storage. Run a scrape first."
        )

    excel_bytes = generate_excel_bytes(data)

    ensure_container_exists(container_name)
    client = get_blob_service_client()

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    blob_name = f"fd_rates_{timestamp}.xlsx"

    blob_client = client.get_blob_client(container=container_name, blob=blob_name)
    blob_client.upload_blob(
        excel_bytes,
        overwrite=True,
        content_settings=ContentSettings(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )

    # Also keep a "latest.xlsx" pointer
    latest_client = client.get_blob_client(container=container_name, blob="latest.xlsx")
    latest_client.upload_blob(
        excel_bytes,
        overwrite=True,
        content_settings=ContentSettings(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )

    return {"excel_url": blob_client.url, "file_name": blob_name}
